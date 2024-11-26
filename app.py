import openpyxl  # Importa a biblioteca para manipulação de arquivos Excel.
from selenium import webdriver  # Importa o Selenium para automação de navegador.
from selenium.webdriver.common.by import By  # Facilita a localização de elementos no DOM.
from time import sleep  # Importa o módulo para pausar a execução do código.

# Carrega a planilha Excel que contém os dados dos clientes.
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
# Seleciona a aba específica da planilha onde os dados estão armazenados.
pagina_clientes = planilha_clientes['Sheet1']

# Inicia uma instância do navegador Chrome para automação.
driver = webdriver.Chrome()
# Acessa o site que será utilizado para consultar os dados.
driver.get('https://consultcpf-devaprender.netlify.app/')

# Percorre cada linha da planilha, começando da segunda linha, para obter os dados dos clientes.
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha  # Atribui os valores da linha a variáveis correspondentes.

    # Localiza o campo de entrada de CPF na página.
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)  # Aguarda um segundo.

    # Limpa o campo de entrada e insere o CPF do cliente.
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)  # Aguarda um segundo.

    # Localiza o botão "Pesquisar" e o clica.
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)  # Aguarda um segundo.

    botao_pesquisar.click()  # Executa o clique no botão.
    sleep(4)  # Aguarda quatro segundos para a resposta do site.

    # Localiza o elemento que contém o status do CPF consultado.
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")

    if status.text == 'em dia':  # Verifica se o status do CPF é "em dia".
        # Obtém informações adicionais, como data e método de pagamento.
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

        # Extrai os valores limpos de data e método de pagamento.
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        # Carrega a planilha de fechamento para registrar os dados processados.
        planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        # Adiciona os dados do cliente à planilha de fechamento com status "em dia".
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

        # Salva a planilha de fechamento atualizada.
        planilha_fechamento.save('planilha_fechamento.xlsx')
    else:  # Se o status não for "em dia".
        # Carrega novamente a planilha de fechamento.
        planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        # Adiciona os dados do cliente à planilha de fechamento com status "pendente".
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

        # Salva a planilha de fechamento atualizada.
        planilha_fechamento.save('planilha_fechamento.xlsx')